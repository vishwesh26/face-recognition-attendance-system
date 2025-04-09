import os
import pandas as pd
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

def mark_attendance(student_id, name, subject, date, time):
    """
    Mark attendance for a student in the attendance record
    
    Args:
        student_id (int): Student ID
        name (str): Student name
        subject (str): Subject name
        date (str): Current date in YYYY-MM-DD format
        time (str): Current time in HH:MM:SS format
        
    Returns:
        bool: True if attendance was marked successfully, False otherwise
    """
    try:
        # File path for the attendance record
        file_path = f"AttendanceRecords/{subject}_attendance.xlsx"
        
        # If file doesn't exist, create it with headers
        if not os.path.exists(file_path):
            _create_attendance_file(file_path, subject)
        
        # Load the workbook
        wb = load_workbook(file_path)
        
        # Check if there's a sheet for the current date, if not create one
        sheet_name = date
        if sheet_name not in wb.sheetnames:
            _create_date_sheet(wb, sheet_name)
        
        # Get the sheet for the current date
        ws = wb[sheet_name]
        
        # Find the next empty row
        row = ws.max_row + 1
        
        # Add the attendance record
        ws.cell(row=row, column=1).value = student_id
        ws.cell(row=row, column=2).value = name
        ws.cell(row=row, column=3).value = time
        ws.cell(row=row, column=4).value = "Present"
        
        # Save the workbook
        wb.save(file_path)
        
        return True
    
    except Exception as e:
        print(f"Error marking attendance: {e}")
        return False

def _create_attendance_file(file_path, subject):
    """
    Create a new attendance file with headers
    
    Args:
        file_path (str): Path to the attendance file
        subject (str): Subject name
    """
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Overview"
    
    # Add title
    ws.merge_cells('A1:D1')
    title_cell = ws['A1']
    title_cell.value = f"{subject} Attendance Records"
    title_cell.font = Font(size=16, bold=True)
    title_cell.alignment = Alignment(horizontal='center')
    
    # Add description
    ws.merge_cells('A2:D2')
    desc_cell = ws['A2']
    desc_cell.value = "This file contains attendance records organized by date"
    desc_cell.alignment = Alignment(horizontal='center')
    
    # Add created date
    ws.merge_cells('A3:D3')
    date_cell = ws['A3']
    date_cell.value = f"Created on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    date_cell.alignment = Alignment(horizontal='center')
    
    # Add instructions
    ws.merge_cells('A5:D5')
    ws['A5'].value = "Each sheet represents a date when attendance was taken"
    ws['A5'].alignment = Alignment(horizontal='center')
    
    # Save the workbook
    wb.save(file_path)

def _create_date_sheet(workbook, sheet_name):
    """
    Create a new sheet for a date in the attendance workbook
    
    Args:
        workbook: The workbook object
        sheet_name (str): Name of the sheet (date)
    """
    # Create a new sheet
    ws = workbook.create_sheet(sheet_name)
    
    # Add headers
    headers = ["ID", "Name", "Time", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15